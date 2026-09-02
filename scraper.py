"""
Veille automatisée Finance / Économie.

Lit des flux RSS, extrait le texte des articles, les classe par thème,
les note de 0 à 10 selon leur pertinence métier via Claude, et produit un
classeur Excel trié, colorié et accompagné d'une feuille de synthèse.

Usage :
    python scraper.py                  # exécution normale
    python scraper.py --sans-ia        # collecte seule, aucun appel API
    python scraper.py --max 3          # 3 articles par source (test rapide)
    python scraper.py --sortie test.xlsx
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import random
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import feedparser
import pandas as pd
import requests
import trafilatura
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import config

# ------------------------------------------------------------------
# Journalisation : console + fichier horodaté dans logs/
# ------------------------------------------------------------------

def configurer_logs() -> logging.Logger:
    Path(config.DOSSIER_LOGS).mkdir(exist_ok=True)
    fichier = Path(config.DOSSIER_LOGS) / f"veille_{datetime.now():%Y-%m-%d}.log"

    logger = logging.getLogger("veille")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter("%(message)s"))

    disque = logging.FileHandler(fichier, encoding="utf-8")
    disque.setLevel(logging.DEBUG)
    disque.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))

    logger.addHandler(console)
    logger.addHandler(disque)
    return logger


log = configurer_logs()

# ------------------------------------------------------------------
# Session HTTP mutualisée, avec réessais automatiques
# ------------------------------------------------------------------

NAVIGATEUR = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"
)


def creer_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": NAVIGATEUR, "Accept-Language": "fr-FR,fr;q=0.9"})
    politique = Retry(
        total=3,
        backoff_factor=1.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD"],
    )
    adaptateur = HTTPAdapter(max_retries=politique, pool_maxsize=config.NB_THREADS_PARALLELES * 2)
    session.mount("https://", adaptateur)
    session.mount("http://", adaptateur)
    return session


SESSION = creer_session()

# ------------------------------------------------------------------
# Structure d'un article
# ------------------------------------------------------------------

@dataclass
class Article:
    date_publication: str
    date_collecte: str
    source: str
    titre: str
    theme: str
    score: int
    pertinent: str
    raison: str
    resume: str
    url: str


# ------------------------------------------------------------------
# Nettoyage des URL (anti-doublons fiable)
# ------------------------------------------------------------------

PARAMS_PARASITES = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "fbclid", "gclid", "xtor", "ref", "ncid", "at_medium", "at_campaign",
}


def normaliser_url(url: str) -> str:
    """Retire les paramètres de tracking pour qu'un même article ne soit pas compté deux fois."""
    try:
        morceaux = urlparse(url)
        params = {k: v for k, v in parse_qs(morceaux.query).items() if k.lower() not in PARAMS_PARASITES}
        propre = morceaux._replace(query=urlencode(params, doseq=True), fragment="")
        return urlunparse(propre).rstrip("/")
    except Exception:
        return url


def resoudre_redirection(url: str) -> str:
    """Les liens Google News pointent vers une redirection : on récupère l'URL réelle."""
    if "news.google.com" not in url:
        return url
    try:
        reponse = SESSION.get(url, timeout=config.TIMEOUT_SECONDES, allow_redirects=True)
        return reponse.url
    except requests.RequestException:
        return url


# ------------------------------------------------------------------
# Classification par mots-clés
# ------------------------------------------------------------------

def classer_article(texte: str) -> str:
    if not texte:
        return config.THEME_PAR_DEFAUT
    minuscules = texte.lower()
    scores = {
        theme: sum(minuscules.count(mot) for mot in infos["mots_cles"])
        for theme, infos in config.THEMES.items()
    }
    meilleur = max(scores, key=scores.get)
    return meilleur if scores[meilleur] > 0 else config.THEME_PAR_DEFAUT


# ------------------------------------------------------------------
# Notation de pertinence par Claude
# ------------------------------------------------------------------

_client_ia = None


def client_ia():
    """Instancie le client Anthropic une seule fois, et seulement si nécessaire."""
    global _client_ia
    if _client_ia is None:
        import anthropic
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise RuntimeError("Variable d'environnement ANTHROPIC_API_KEY absente.")
        _client_ia = anthropic.Anthropic()
    return _client_ia


INSTRUCTION_IA = """Tu évalues la pertinence d'un article de presse pour un professionnel.

Son métier : {profil}

Titre : {titre}
Extrait : {extrait}

Réponds UNIQUEMENT par un objet JSON valide, sans texte autour, sans balises Markdown :
{{"score": <entier de 0 à 10>, "raison": "<une phrase de 15 mots maximum>"}}

Barème : 0-3 hors sujet, 4-6 tangent, 7-8 utile, 9-10 à lire en priorité."""


def evaluer_pertinence(titre: str, texte: str) -> tuple[int, str]:
    """Retourne (score sur 10, justification). Score -1 si l'évaluation a échoué."""
    invite = INSTRUCTION_IA.format(
        profil=config.PROFIL_METIER, titre=titre, extrait=(texte or "")[:2000]
    )

    for tentative in range(3):
        try:
            reponse = client_ia().messages.create(
                model=config.MODELE_IA,
                max_tokens=150,
                messages=[{"role": "user", "content": invite}],
            )
            brut = reponse.content[0].text.strip().removeprefix("```json").removeprefix("```").removesuffix("```")
            donnees = json.loads(brut)
            score = max(0, min(10, int(donnees.get("score", 0))))
            return score, str(donnees.get("raison", ""))[:200]
        except json.JSONDecodeError:
            log.debug("Réponse IA non exploitable pour : %s", titre[:60])
            return -1, "Réponse IA illisible"
        except Exception as erreur:
            attente = (2 ** tentative) + random.random()
            log.debug("Appel IA échoué (essai %d) : %s — nouvelle tentative dans %.1fs",
                      tentative + 1, erreur, attente)
            time.sleep(attente)

    return -1, "API indisponible"


# ------------------------------------------------------------------
# Lecture des flux
# ------------------------------------------------------------------

def date_publication(entree) -> datetime | None:
    for champ in ("published_parsed", "updated_parsed"):
        valeur = entree.get(champ)
        if valeur:
            try:
                return datetime(*valeur[:6], tzinfo=timezone.utc)
            except (TypeError, ValueError):
                continue
    return None


def lire_flux(nom: str, url_flux: str, deja_vues: set[str]) -> list[dict]:
    log.info("Lecture du flux : %s", nom)
    try:
        reponse = SESSION.get(url_flux, timeout=config.TIMEOUT_SECONDES)
        reponse.raise_for_status()
        flux = feedparser.parse(reponse.content)
    except requests.RequestException as erreur:
        log.warning("  %s injoignable : %s", nom, erreur)
        return []

    if not flux.entries:
        log.warning("  %s : aucun article dans le flux", nom)
        return []

    limite = (
        datetime.now(timezone.utc) - timedelta(days=config.AGE_MAX_JOURS)
        if config.AGE_MAX_JOURS else None
    )

    retenues = []
    for entree in flux.entries[: config.NB_ARTICLES_PAR_SOURCE]:
        lien = entree.get("link", "")
        if not lien:
            continue

        publiee = date_publication(entree)
        if limite and publiee and publiee < limite:
            continue

        cle = normaliser_url(lien)
        if cle in deja_vues:
            continue
        deja_vues.add(cle)

        retenues.append({
            "source": nom,
            "titre": entree.get("title", "Sans titre").strip(),
            "url": lien,
            "url_propre": cle,
            "resume_rss": entree.get("summary", ""),
            "publiee": publiee.strftime("%Y-%m-%d %H:%M") if publiee else "",
        })

    log.info("  %d nouvel(s) article(s) retenu(s)", len(retenues))
    return retenues


def extraire_texte(url: str, secours: str) -> str:
    try:
        page = SESSION.get(url, timeout=config.TIMEOUT_SECONDES)
        texte = trafilatura.extract(page.text, include_comments=False, include_tables=False)
        if texte and len(texte) > 200:
            return texte
    except Exception as erreur:
        log.debug("Extraction impossible sur %s : %s", url[:80], erreur)
    return trafilatura.html2txt(secours) if secours else ""


def traiter(entree: dict, avec_ia: bool) -> Article:
    url_reelle = resoudre_redirection(entree["url"])
    texte = extraire_texte(url_reelle, entree["resume_rss"])
    theme = classer_article(texte)

    if avec_ia:
        score, raison = evaluer_pertinence(entree["titre"], texte)
    else:
        score, raison = -1, "IA désactivée"

    if score < 0:
        pertinent = "?"
    elif score >= config.SEUIL_PERTINENCE:
        pertinent = "Oui"
    else:
        pertinent = "Non"

    log.info("  [%s] %-22s %s", f"{score:2d}" if score >= 0 else " -", theme[:22], entree["titre"][:55])

    return Article(
        date_publication=entree["publiee"],
        date_collecte=datetime.now().strftime("%Y-%m-%d %H:%M"),
        source=entree["source"],
        titre=entree["titre"],
        theme=theme,
        score=score,
        pertinent=pertinent,
        raison=raison,
        resume=" ".join(texte.split())[:300],
        url=url_reelle,
    )


# ------------------------------------------------------------------
# Historique
# ------------------------------------------------------------------

COLONNES = ["date_publication", "date_collecte", "source", "titre", "theme",
            "score", "pertinent", "raison", "resume", "url"]


def charger_historique(chemin: str) -> tuple[pd.DataFrame, set[str]]:
    if not Path(chemin).exists():
        return pd.DataFrame(columns=COLONNES), set()
    try:
        df = pd.read_excel(chemin, sheet_name=0)
        for colonne in COLONNES:
            if colonne not in df.columns:
                df[colonne] = ""
        df = df[COLONNES]
        log.info("Historique chargé : %d article(s) déjà connus.", len(df))
        return df, {normaliser_url(u) for u in df["url"].dropna().astype(str)}
    except Exception as erreur:
        log.warning("Historique illisible (%s) — on repart à zéro.", erreur)
        return pd.DataFrame(columns=COLONNES), set()


# ------------------------------------------------------------------
# Rendu Excel
# ------------------------------------------------------------------

GRIS_FONCE = "404040"
BORDURE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def trier(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_ordre"] = df["pertinent"].map({"Oui": 0, "Non": 1, "?": 2}).fillna(2)
    df["_date"] = pd.to_datetime(df["date_publication"], errors="coerce")
    df["_score"] = pd.to_numeric(df["score"], errors="coerce").fillna(-1)
    return (df.sort_values(["_ordre", "_score", "theme", "_date"],
                           ascending=[True, False, True, False])
              .drop(columns=["_ordre", "_date", "_score"])
              .reset_index(drop=True))


def feuille_synthese(writer, df: pd.DataFrame) -> None:
    scores = pd.to_numeric(df["score"], errors="coerce")

    par_theme = (df.groupby("theme")
                   .agg(articles=("titre", "count"),
                        pertinents=("pertinent", lambda s: (s == "Oui").sum()))
                   .reset_index()
                   .sort_values("articles", ascending=False))

    par_source = (df.groupby("source")
                    .agg(articles=("titre", "count"),
                         pertinents=("pertinent", lambda s: (s == "Oui").sum()))
                    .reset_index()
                    .sort_values("articles", ascending=False))

    global_ = pd.DataFrame({
        "Indicateur": ["Articles au total", "Articles pertinents", "Taux de pertinence",
                       "Score moyen", "Sources actives", "Dernière collecte"],
        "Valeur": [
            len(df),
            int((df["pertinent"] == "Oui").sum()),
            f"{(df['pertinent'] == 'Oui').mean():.0%}" if len(df) else "0%",
            f"{scores[scores >= 0].mean():.1f} / 10" if (scores >= 0).any() else "n/a",
            df["source"].nunique(),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    })

    global_.to_excel(writer, sheet_name="Synthèse", index=False, startrow=0)
    par_theme.to_excel(writer, sheet_name="Synthèse", index=False, startrow=len(global_) + 3)
    par_source.to_excel(writer, sheet_name="Synthèse", index=False,
                        startrow=len(global_) + len(par_theme) + 6)


def styliser(chemin: str, df: pd.DataFrame) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(chemin)
    ws = wb["Articles"]
    colonnes = {nom: i + 1 for i, nom in enumerate(df.columns)}

    for cellule in ws[1]:
        cellule.font = Font(bold=True, color="FFFFFF", size=11)
        cellule.fill = PatternFill("solid", start_color=GRIS_FONCE)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for i, ligne in enumerate(df.itertuples(index=False), start=2):
        couleur = config.THEMES.get(ligne.theme, {}).get("couleur", config.COULEUR_DEFAUT)
        remplissage = PatternFill("solid", start_color=couleur)
        gras = ligne.pertinent == "Oui"

        for col in range(1, len(df.columns) + 1):
            cellule = ws.cell(row=i, column=col)
            cellule.fill = remplissage
            cellule.border = BORDURE
            cellule.alignment = Alignment(vertical="top", wrap_text=(col == colonnes["resume"]))
            if gras:
                cellule.font = Font(bold=True)

        cellule_titre = ws.cell(row=i, column=colonnes["titre"])
        if isinstance(ligne.url, str) and ligne.url.startswith("http"):
            cellule_titre.hyperlink = ligne.url
            cellule_titre.font = Font(bold=gras, color="0563C1", underline="single")

    largeurs = {"date_publication": 17, "date_collecte": 17, "source": 20, "titre": 60,
                "theme": 24, "score": 7, "pertinent": 10, "raison": 42, "resume": 55, "url": 45}
    for nom, index in colonnes.items():
        ws.column_dimensions[get_column_letter(index)].width = largeurs.get(nom, 18)

    if len(df):
        lettre = get_column_letter(colonnes["score"])
        ws.conditional_formatting.add(
            f"{lettre}2:{lettre}{len(df) + 1}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=5, mid_color="FFEB84",
                           end_type="num", end_value=10, end_color="63BE7B"),
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    synthese = wb["Synthèse"]
    for ligne in synthese.iter_rows():
        for cellule in ligne:
            if cellule.value is not None and cellule.row == 1 or (
                isinstance(cellule.value, str) and cellule.value in
                ("theme", "source", "articles", "pertinents", "Indicateur", "Valeur")
            ):
                cellule.font = Font(bold=True, color="FFFFFF")
                cellule.fill = PatternFill("solid", start_color=GRIS_FONCE)
    for lettre in ("A", "B", "C"):
        synthese.column_dimensions[lettre].width = 26

    wb.save(chemin)


def ecrire_excel(df: pd.DataFrame, chemin: str) -> str:
    df = trier(df)
    cible = chemin
    try:
        with pd.ExcelWriter(cible, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Articles", index=False)
            feuille_synthese(writer, df)
    except PermissionError:
        cible = f"{Path(chemin).stem}_{datetime.now():%H%M%S}.xlsx"
        log.warning("%s est ouvert dans Excel — écriture dans %s à la place.", chemin, cible)
        with pd.ExcelWriter(cible, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Articles", index=False)
            feuille_synthese(writer, df)

    styliser(cible, df)
    return cible


# ------------------------------------------------------------------
# Programme principal
# ------------------------------------------------------------------

def analyser_arguments():
    parseur = argparse.ArgumentParser(description="Veille automatisée Finance / Économie")
    parseur.add_argument("--sans-ia", action="store_true", help="désactive la notation par Claude")
    parseur.add_argument("--max", type=int, metavar="N", help="articles par source (test rapide)")
    parseur.add_argument("--sortie", metavar="FICHIER", help="nom du fichier Excel produit")
    return parseur.parse_args()


def main() -> int:
    debut = time.time()
    args = analyser_arguments()

    if args.max:
        config.NB_ARTICLES_PAR_SOURCE = args.max
    sortie = args.sortie or config.FICHIER_SORTIE

    avec_ia = config.IA_ACTIVE and not args.sans_ia
    if avec_ia and not os.environ.get("ANTHROPIC_API_KEY"):
        log.warning("ANTHROPIC_API_KEY introuvable — passage en mode collecte seule.\n"
                    "         Définis la clé, ou lance : python scraper.py --sans-ia")
        avec_ia = False

    log.info("=" * 62)
    log.info("VEILLE  %s  |  %d sources  |  IA %s",
             datetime.now().strftime("%d/%m/%Y %H:%M"), len(config.SOURCES),
             "activée" if avec_ia else "désactivée")
    log.info("=" * 62)

    historique, deja_vues = charger_historique(sortie)

    a_traiter = []
    for nom, url_flux in config.SOURCES.items():
        a_traiter.extend(lire_flux(nom, url_flux, deja_vues))

    if not a_traiter:
        log.info("\nAucun nouvel article depuis la dernière exécution.")
        if len(historique):
            ecrire_excel(historique, sortie)
        return 0

    log.info("\n%d article(s) à analyser (%d en parallèle)...\n",
             len(a_traiter), config.NB_THREADS_PARALLELES)

    articles = []
    with ThreadPoolExecutor(max_workers=config.NB_THREADS_PARALLELES) as pool:
        futurs = {pool.submit(traiter, e, avec_ia): e for e in a_traiter}
        for futur in as_completed(futurs):
            try:
                articles.append(asdict(futur.result()))
            except Exception as erreur:
                log.error("Article abandonné (%s) : %s", futurs[futur]["titre"][:50], erreur)

    complet = pd.concat([historique, pd.DataFrame(articles)], ignore_index=True)
    fichier = ecrire_excel(complet, sortie)

    pertinents = sum(1 for a in articles if a["pertinent"] == "Oui")
    log.info("\n" + "=" * 62)
    log.info("%d article(s) ajouté(s), dont %d pertinent(s). Total : %d",
             len(articles), pertinents, len(complet))
    log.info("Fichier : %s   |   Durée : %.0f s", fichier, time.time() - debut)
    log.info("=" * 62)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        log.info("\nInterrompu par l'utilisateur.")
        sys.exit(130)
    except Exception as erreur:
        log.exception("Erreur fatale : %s", erreur)
        sys.exit(1)
